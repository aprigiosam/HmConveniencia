"""
Serviço de exportação de relatórios em múltiplos formatos
Suporta PDF, Excel (XLSX) e CSV
"""

from io import BytesIO
from decimal import Decimal
from datetime import datetime
from typing import Dict, Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import csv


class ExportService:
    """Serviço para exportação de relatórios"""

    @staticmethod
    def exportar_relatorio_pdf(relatorio: Dict[str, Any]) -> BytesIO:
        """
        Exporta relatório para PDF usando ReportLab

        Args:
            relatorio: Dados do relatório

        Returns:
            BytesIO com o PDF gerado
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=15*mm,
            leftMargin=15*mm,
            topMargin=15*mm,
            bottomMargin=15*mm
        )

        # Estilos
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1e40af'),
            spaceAfter=20,
            alignment=TA_CENTER,
        )

        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1f2937'),
            spaceAfter=12,
            spaceBefore=12,
        )

        normal_style = styles['Normal']

        # Elementos do documento
        elements = []

        # Cabeçalho
        sessao = relatorio['sessao']
        elements.append(Paragraph(f"Relatório {sessao['tipo']} - {sessao['codigo']}", title_style))
        elements.append(Paragraph(f"{sessao['loja']['nome']}", normal_style))
        if sessao['loja'].get('cnpj'):
            elements.append(Paragraph(f"CNPJ: {sessao['loja']['cnpj']}", normal_style))
        elements.append(Paragraph(
            f"Data: {datetime.fromisoformat(sessao['data_relatorio'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')}",
            normal_style
        ))
        elements.append(Spacer(1, 12))

        # Informações da Sessão
        elements.append(Paragraph("Informações da Sessão", heading_style))

        info_data = [
            ['Responsável:', sessao['responsavel']],
            ['Status:', sessao['status']],
            ['Abertura:', datetime.fromisoformat(sessao['aberta_em'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')],
        ]

        if sessao.get('fechada_em'):
            info_data.append(['Fechamento:', datetime.fromisoformat(sessao['fechada_em'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')])

        info_table = Table(info_data, colWidths=[40*mm, 120*mm])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 12))

        # Resumo de Vendas
        resumo = relatorio['resumo']
        elements.append(Paragraph("Resumo de Vendas", heading_style))

        resumo_data = [
            ['Total de Vendas:', f"R$ {resumo['total_vendas']:.2f}"],
            ['Quantidade de Vendas:', str(resumo['quantidade_vendas'])],
            ['Ticket Médio:', f"R$ {resumo['ticket_medio']:.2f}"],
            ['Total Descontos:', f"R$ {resumo['total_descontos']:.2f}"],
            ['Vendas Canceladas:', str(resumo['vendas_canceladas'])],
        ]

        resumo_table = Table(resumo_data, colWidths=[50*mm, 110*mm])
        resumo_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#dcfce7')),
            ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#dcfce7')),
        ]))
        elements.append(resumo_table)
        elements.append(Spacer(1, 12))

        # Movimentação de Caixa
        caixa = relatorio['caixa']
        elements.append(Paragraph("Movimentação de Caixa", heading_style))

        caixa_data = [
            ['Saldo Inicial:', f"R$ {caixa['saldo_inicial']:.2f}"],
            ['+ Vendas:', f"R$ {caixa['total_vendas']:.2f}"],
            ['- Sangrias:', f"R$ {caixa['total_sangrias']:.2f}"],
            ['+ Reforços:', f"R$ {caixa['total_reforcos']:.2f}"],
            ['= Saldo Teórico:', f"R$ {caixa['saldo_teorico']:.2f}"],
            ['Saldo Real:', f"R$ {caixa['saldo_real']:.2f}"],
            ['Diferença:', f"R$ {caixa['diferenca']:+.2f}"],
        ]

        caixa_table = Table(caixa_data, colWidths=[50*mm, 110*mm])
        caixa_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LINEABOVE', (0, 4), (1, 4), 1, colors.grey),
            ('LINEABOVE', (0, 5), (1, 5), 1, colors.grey),
            ('FONTNAME', (0, 6), (1, 6), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 6), (1, 6), colors.HexColor('#fef3c7') if abs(caixa['diferenca']) > 1 else colors.HexColor('#dcfce7')),
        ]))
        elements.append(caixa_table)
        elements.append(Spacer(1, 12))

        # Formas de Pagamento
        if relatorio.get('formas_pagamento'):
            elements.append(Paragraph("Formas de Pagamento", heading_style))

            pagamento_data = [['Forma', 'Quantidade', 'Valor']]
            for fp in relatorio['formas_pagamento']:
                pagamento_data.append([
                    fp['nome'],
                    str(fp['quantidade']),
                    f"R$ {fp['valor']:.2f}"
                ])

            pagamento_table = Table(pagamento_data, colWidths=[80*mm, 40*mm, 40*mm])
            pagamento_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
            ]))
            elements.append(pagamento_table)
            elements.append(Spacer(1, 12))

        # Top Produtos
        if relatorio.get('top_produtos') and len(relatorio['top_produtos']) > 0:
            elements.append(Paragraph("Top Produtos Mais Vendidos", heading_style))

            produtos_data = [['Produto', 'SKU', 'Qtd', 'Valor']]
            for produto in relatorio['top_produtos'][:10]:
                produtos_data.append([
                    produto['nome'][:40],
                    produto['sku'],
                    f"{produto['quantidade']:.0f}",
                    f"R$ {produto['valor']:.2f}"
                ])

            produtos_table = Table(produtos_data, colWidths=[80*mm, 30*mm, 25*mm, 25*mm])
            produtos_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(produtos_table)

        # Rodapé
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}",
            ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        ))

        # Gera o PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def exportar_relatorio_excel(relatorio: Dict[str, Any]) -> BytesIO:
        """
        Exporta relatório para Excel usando openpyxl

        Args:
            relatorio: Dados do relatório

        Returns:
            BytesIO com o arquivo Excel gerado
        """
        buffer = BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório"

        # Estilos
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")

        subheader_font = Font(bold=True, size=12)
        subheader_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

        bold_font = Font(bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        row = 1
        sessao = relatorio['sessao']

        # Título
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = f"Relatório {sessao['tipo']} - {sessao['codigo']}"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 1

        # Informações da Loja
        ws[f'A{row}'] = sessao['loja']['nome']
        ws[f'A{row}'].font = bold_font
        row += 1

        if sessao['loja'].get('cnpj'):
            ws[f'A{row}'] = f"CNPJ: {sessao['loja']['cnpj']}"
            row += 1

        ws[f'A{row}'] = f"Data: {datetime.fromisoformat(sessao['data_relatorio'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')}"
        row += 2

        # Informações da Sessão
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "Informações da Sessão"
        cell.font = subheader_font
        cell.fill = subheader_fill
        row += 1

        ws[f'A{row}'] = "Responsável:"
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = sessao['responsavel']
        row += 1

        ws[f'A{row}'] = "Status:"
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = sessao['status']
        row += 1

        ws[f'A{row}'] = "Abertura:"
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = datetime.fromisoformat(sessao['aberta_em'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
        row += 1

        if sessao.get('fechada_em'):
            ws[f'A{row}'] = "Fechamento:"
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = datetime.fromisoformat(sessao['fechada_em'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')
            row += 1

        row += 1

        # Resumo de Vendas
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "Resumo de Vendas"
        cell.font = subheader_font
        cell.fill = subheader_fill
        row += 1

        resumo = relatorio['resumo']
        resumo_items = [
            ('Total de Vendas:', f"R$ {resumo['total_vendas']:.2f}"),
            ('Quantidade de Vendas:', resumo['quantidade_vendas']),
            ('Ticket Médio:', f"R$ {resumo['ticket_medio']:.2f}"),
            ('Total Descontos:', f"R$ {resumo['total_descontos']:.2f}"),
            ('Vendas Canceladas:', resumo['vendas_canceladas']),
        ]

        for label, value in resumo_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Movimentação de Caixa
        ws.merge_cells(f'A{row}:B{row}')
        cell = ws[f'A{row}']
        cell.value = "Movimentação de Caixa"
        cell.font = subheader_font
        cell.fill = subheader_fill
        row += 1

        caixa = relatorio['caixa']
        caixa_items = [
            ('Saldo Inicial:', f"R$ {caixa['saldo_inicial']:.2f}"),
            ('+ Vendas:', f"R$ {caixa['total_vendas']:.2f}"),
            ('- Sangrias:', f"R$ {caixa['total_sangrias']:.2f}"),
            ('+ Reforços:', f"R$ {caixa['total_reforcos']:.2f}"),
            ('= Saldo Teórico:', f"R$ {caixa['saldo_teorico']:.2f}"),
            ('Saldo Real:', f"R$ {caixa['saldo_real']:.2f}"),
            ('Diferença:', f"R$ {caixa['diferenca']:+.2f}"),
        ]

        for label, value in caixa_items:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'B{row}'] = value
            row += 1

        row += 1

        # Formas de Pagamento
        if relatorio.get('formas_pagamento'):
            ws.merge_cells(f'A{row}:C{row}')
            cell = ws[f'A{row}']
            cell.value = "Formas de Pagamento"
            cell.font = subheader_font
            cell.fill = subheader_fill
            row += 1

            # Cabeçalho
            headers = ['Forma', 'Quantidade', 'Valor']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = bold_font
                cell.border = border
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            row += 1

            # Dados
            for fp in relatorio['formas_pagamento']:
                ws[f'A{row}'] = fp['nome']
                ws[f'B{row}'] = fp['quantidade']
                ws[f'C{row}'] = f"R$ {fp['valor']:.2f}"

                for col in range(1, 4):
                    ws.cell(row=row, column=col).border = border
                row += 1

            row += 1

        # Top Produtos
        if relatorio.get('top_produtos') and len(relatorio['top_produtos']) > 0:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "Top Produtos Mais Vendidos"
            cell.font = subheader_font
            cell.fill = subheader_fill
            row += 1

            # Cabeçalho
            headers = ['Produto', 'SKU', 'Quantidade', 'Valor']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.font = bold_font
                cell.border = border
                cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
            row += 1

            # Dados
            for produto in relatorio['top_produtos'][:10]:
                ws[f'A{row}'] = produto['nome']
                ws[f'B{row}'] = produto['sku']
                ws[f'C{row}'] = produto['quantidade']
                ws[f'D{row}'] = f"R$ {produto['valor']:.2f}"

                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                row += 1

        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15

        # Salvar
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @staticmethod
    def exportar_relatorio_csv(relatorio: Dict[str, Any]) -> BytesIO:
        """
        Exporta relatório para CSV

        Args:
            relatorio: Dados do relatório

        Returns:
            BytesIO com o arquivo CSV gerado
        """
        buffer = BytesIO()

        # CSV precisa de text mode
        import io
        text_buffer = io.StringIO()
        writer = csv.writer(text_buffer, delimiter=';')

        sessao = relatorio['sessao']

        # Cabeçalho
        writer.writerow([f"Relatório {sessao['tipo']} - {sessao['codigo']}"])
        writer.writerow([sessao['loja']['nome']])
        if sessao['loja'].get('cnpj'):
            writer.writerow([f"CNPJ: {sessao['loja']['cnpj']}"])
        writer.writerow([f"Data: {datetime.fromisoformat(sessao['data_relatorio'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')}"])
        writer.writerow([])

        # Informações da Sessão
        writer.writerow(['Informações da Sessão'])
        writer.writerow(['Responsável', sessao['responsavel']])
        writer.writerow(['Status', sessao['status']])
        writer.writerow(['Abertura', datetime.fromisoformat(sessao['aberta_em'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')])
        if sessao.get('fechada_em'):
            writer.writerow(['Fechamento', datetime.fromisoformat(sessao['fechada_em'].replace('Z', '+00:00')).strftime('%d/%m/%Y %H:%M')])
        writer.writerow([])

        # Resumo
        resumo = relatorio['resumo']
        writer.writerow(['Resumo de Vendas'])
        writer.writerow(['Total de Vendas', f"R$ {resumo['total_vendas']:.2f}"])
        writer.writerow(['Quantidade de Vendas', resumo['quantidade_vendas']])
        writer.writerow(['Ticket Médio', f"R$ {resumo['ticket_medio']:.2f}"])
        writer.writerow(['Total Descontos', f"R$ {resumo['total_descontos']:.2f}"])
        writer.writerow(['Vendas Canceladas', resumo['vendas_canceladas']])
        writer.writerow([])

        # Caixa
        caixa = relatorio['caixa']
        writer.writerow(['Movimentação de Caixa'])
        writer.writerow(['Saldo Inicial', f"R$ {caixa['saldo_inicial']:.2f}"])
        writer.writerow(['+ Vendas', f"R$ {caixa['total_vendas']:.2f}"])
        writer.writerow(['- Sangrias', f"R$ {caixa['total_sangrias']:.2f}"])
        writer.writerow(['+ Reforços', f"R$ {caixa['total_reforcos']:.2f}"])
        writer.writerow(['= Saldo Teórico', f"R$ {caixa['saldo_teorico']:.2f}"])
        writer.writerow(['Saldo Real', f"R$ {caixa['saldo_real']:.2f}"])
        writer.writerow(['Diferença', f"R$ {caixa['diferenca']:+.2f}"])
        writer.writerow([])

        # Formas de Pagamento
        if relatorio.get('formas_pagamento'):
            writer.writerow(['Formas de Pagamento'])
            writer.writerow(['Forma', 'Quantidade', 'Valor'])
            for fp in relatorio['formas_pagamento']:
                writer.writerow([fp['nome'], fp['quantidade'], f"R$ {fp['valor']:.2f}"])
            writer.writerow([])

        # Top Produtos
        if relatorio.get('top_produtos') and len(relatorio['top_produtos']) > 0:
            writer.writerow(['Top Produtos Mais Vendidos'])
            writer.writerow(['Produto', 'SKU', 'Quantidade', 'Valor'])
            for produto in relatorio['top_produtos'][:10]:
                writer.writerow([
                    produto['nome'],
                    produto['sku'],
                    produto['quantidade'],
                    f"R$ {produto['valor']:.2f}"
                ])

        # Converter para bytes
        buffer.write(text_buffer.getvalue().encode('utf-8-sig'))  # BOM para Excel abrir corretamente
        buffer.seek(0)
        return buffer